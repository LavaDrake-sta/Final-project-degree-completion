"""
Excel Redactor - השחרת קבצי Excel
פרויקט גמר - זיהוי מידע אישי רגיש

מודול להשחרה אקטיבית של PII מקבצי Excel
"""

import openpyxl
from openpyxl.styles import PatternFill, Font
import io
import logging
import random
from typing import List, Union, Optional

class ExcelRedactor:
    """
    מחלקה האחראית על השחרת מידע רגיש מקבצי אקסל.
    המחלקה מקבלת קובץ ורשימת טקסטים להסרה, משנה את הטקסט באופן בלתי הפיך ל-0 ול-1
    ומייצרת קובץ אקסל חדש ובטוח.
    """

    def __init__(self):
        self._setup_logging()

    def _setup_logging(self):
        logging.basicConfig(level=logging.INFO)
        self.logger = logging.getLogger(__name__)

    def _mask_text_binary(self, original_text: str) -> str:
        """מחליף כל תו בטקסט הרגיש ב-0 או 1 אקראיים כדי למנוע שחזור"""
        return "".join(random.choice(['0', '1']) for _ in original_text)

    def redact_excel(self, excel_data: Union[str, bytes], pii_texts: List[str], output_path: Optional[str] = None) -> Union[bytes, str, bool]:
        """
        השחרת נתונים רגישים מתוך קובץ Excel.
        - excel_data: נתיב לקובץ מקור או נתוני bytes
        - pii_texts: רשימה של מחרוזות (הטקסט של המידע הרגיש) שיש להשחיר
        - output_path: נתיב לשמירת הקובץ. אם None, יוחזרו bytes.
        """
        try:
            if isinstance(excel_data, str):
                wb = openpyxl.load_workbook(excel_data)
            else:
                wb = openpyxl.load_workbook(io.BytesIO(excel_data))
                
            black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
            black_font = Font(color="000000")

            self.logger.info(f"🔒 מתחיל השחרת קובץ Excel. מספר מחרוזות PII להשחרה: {len(pii_texts)}")
            redact_count = 0

            # מעבר על כל הגיליונות והתאים
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str):
                            original_value = cell.value
                            new_value = original_value
                            modified = False
                            
                            # חיפוש תבניות PII בטקסט שבתא
                            for pii in pii_texts:
                                if pii and pii in new_value:
                                    # יוצר מסכת 0 ו-1 באורך הטקסט המקורי
                                    masked_pii = self._mask_text_binary(pii)
                                    # החלפת הטקסט בפועל מונעת כל אפשרות לשחזור המקורי
                                    new_value = new_value.replace(pii, masked_pii)
                                    modified = True
                                    redact_count += 1
                                    
                            if modified:
                                cell.value = new_value
                                cell.fill = black_fill
                                cell.font = black_font

            self.logger.info(f"✅ סיום השחרה. {redact_count} חלקי מידע הופכו למחרוזת בינארית (0 ו-1).")

            # שמירת התוצאה - החזרת הקובץ למשתמש להורדה
            if output_path is not None:
                wb.save(output_path)
                return output_path
            else:
                output = io.BytesIO()
                wb.save(output)
                return output.getvalue()
                
        except Exception as e:
            self.logger.error(f"❌ שגיאה בהשחרת Excel: {e}")
            return False