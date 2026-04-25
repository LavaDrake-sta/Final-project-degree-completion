import sys
import os
from openpyxl import Workbook

# Add src to path
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), 'src')))

from redactors.excel_redactor import ExcelRedactor

def test_excel_redactor():
    print("Testing Excel Redactor...")
    
    # Create a dummy excel file
    wb = Workbook()
    ws = wb.active
    ws['A1'] = "שם: ישראל ישראלי"
    ws['A2'] = "ת.ז: 123456789"
    ws['B1'] = "טלפון: 052-1234567"
    ws['B2'] = "כתובת: רחוב הרצל 15 תל אביב"
    
    dummy_path = "dummy_test.xlsx"
    wb.save(dummy_path)
    
    # Initialize redactor
    redactor = ExcelRedactor()
    
    # The PII strings we pretend the system detected
    pii_texts = ["123456789", "052-1234567", "ישראל ישראלי"]
    
    output_path = "dummy_test_redacted.xlsx"
    
    result = redactor.redact_excel(dummy_path, pii_texts, output_path)
    
    if result:
        print(f"Redaction successful! Saved to {result}")
    else:
        print("Redaction failed!")
        
    # Clean up (optional)
    if os.path.exists(dummy_path):
        os.remove(dummy_path)

if __name__ == "__main__":
    test_excel_redactor()
